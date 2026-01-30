from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from .models import Item, StockBalance, AuditLog, User


def export_outlet_xlsx(db: Session, outlet_id: int, file_path: str) -> str:
    wb = Workbook()

    # -------- Sheet 1: Inventory --------
    ws = wb.active
    ws.title = "Inventory"

    ws.append(["GeneratedAt(UTC)", datetime.utcnow().isoformat(timespec="seconds")])
    ws.append(["OutletID", outlet_id])
    ws.append([])
    ws.append(
        ["ItemID", "Name", "Unit", "Quantity", "CreatedAt(UTC)", "UpdatedAt(UTC)"]
    )

    items = db.scalars(
        select(Item)
        .where(Item.outlet_id == outlet_id, Item.is_active == True)
        .order_by(Item.name.asc())
    ).all()
    if items:
        item_ids = [it.id for it in items]
        bals = db.scalars(
            select(StockBalance).where(
                StockBalance.outlet_id == outlet_id, StockBalance.item_id.in_(item_ids)
            )
        ).all()
        bmap = {b.item_id: b for b in bals}

        for it in items:
            qty = bmap.get(it.id).quantity if bmap.get(it.id) else 0
            created = getattr(it, "created_at", None)
            updated = getattr(it, "updated_at", None)
            ws.append(
                [
                    it.id,
                    it.name,
                    it.unit,
                    float(qty) if qty is not None else 0,
                    created.isoformat(timespec="seconds") if created else None,
                    updated.isoformat(timespec="seconds") if updated else None,
                ]
            )

    # autosize columns (простенько)
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # -------- Sheet 2: Audit --------
    ws2 = wb.create_sheet("Audit")
    ws2.append(
        [
            "Time(UTC)",
            "User",
            "Action",
            "EntityType",
            "EntityID",
            "OutletID",
            "GroupID",
            "Details",
        ]
    )

    # подмешаем имя пользователя
    # (можно оптимизировать join'ом, но для прототипа ок)
    logs = db.scalars(
        select(AuditLog)
        .where(AuditLog.outlet_id == outlet_id)
        .order_by(AuditLog.created_at.asc())
    ).all()

    user_ids = sorted({l.user_id for l in logs})
    users = {}
    if user_ids:
        us = db.scalars(select(User).where(User.id.in_(user_ids))).all()
        users = {u.id: (u.name or str(u.tg_user_id)) for u in us}

    for l in logs:
        ws2.append(
            [
                l.created_at.isoformat(timespec="seconds"),
                users.get(l.user_id, str(l.user_id)),
                l.action.value if hasattr(l.action, "value") else str(l.action),
                l.entity_type,
                l.entity_id,
                l.outlet_id,
                l.group_id,
                l.details,
            ]
        )

    for col in range(1, 9):
        ws2.column_dimensions[get_column_letter(col)].width = 22

    Path(file_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(file_path)
    return file_path
